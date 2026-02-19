#!/usr/bin/env python3
"""
Generate Instrument Index Excel from database.

Creates a 19-column Excel file matching standard instrument index format.

Usage:
    python generate_instrument_index.py --database database.yaml
    python generate_instrument_index.py -d database.yaml -o instrument-index.xlsx
"""

import argparse
import sys
from pathlib import Path

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_yaml(path: Path) -> dict:
    """Load YAML file."""
    with open(path) as f:
        return yaml.safe_load(f)


# Column definitions for Instrument Index
COLUMNS = [
    ("ITEM", 8),
    ("TAG NUMBER", 15),
    ("SERVICE DESCRIPTION", 35),
    ("P&ID DRG NO", 12),
    ("EQUIPMENT LOCATION", 18),
    ("LOCATION", 10),
    ("MAKE", 15),
    ("TYPE OF INSTRUMENT", 25),
    ("SIGNAL TYPE", 12),
    ("ENGG UNITS", 10),
    ("LOW RANGE", 10),
    ("MAX RANGE", 10),
    ("LO-LO ALARM", 10),
    ("LOW ALARM", 10),
    ("HI ALARM", 10),
    ("HI-HI ALARM", 10),
    ("PLC 4mA Value", 12),
    ("PLC 20mA Value", 12),
    ("REMARKS", 30),
]


def create_workbook(instruments: list, project_id: str, revision: dict) -> Workbook:
    """
    Create Instrument Index workbook.

    Args:
        instruments: List of instrument entries
        project_id: Project identifier
        revision: Revision info

    Returns:
        openpyxl Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Instrument Index"

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title row
    ws.merge_cells("A1:S1")
    ws["A1"] = f"INSTRUMENT INDEX - {project_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Revision row
    ws.merge_cells("A2:S2")
    ws["A2"] = f"Revision: {revision.get('number', 'A')} | Date: {revision.get('date', '')} | By: {revision.get('by', '')}"
    ws["A2"].alignment = center_align

    # Header row (row 4)
    header_row = 4
    for col, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    for item_no, inst in enumerate(instruments, 1):
        row = header_row + item_no
        tag = inst.get("tag", {}) if isinstance(inst.get("tag"), dict) else {}
        # Map to actual YAML schema fields (flat structure, not nested objects)
        # instrument_type, range, output_signal, area etc. are top-level fields

        # Primary signal type
        signal_type = inst.get("primary_signal_type", "")
        # Derive from io_signals if not set
        if not signal_type:
            io_signals = inst.get("io_signals", [])
            if io_signals:
                first_sig = io_signals[0]
                signal_type = first_sig.get("signal_type", first_sig.get("io_type", ""))

        # Range fields — check top-level and nested
        range_val = inst.get("range", "")
        range_unit = inst.get("range_unit", "")
        range_min = inst.get("range_min", "")
        range_max = inst.get("range_max", "")

        # Parse "0-100 degC" style range string into min/max/unit
        if range_val and not range_min and not range_max:
            import re as _re
            range_match = _re.match(r"([\d.]+)\s*[-–]\s*([\d.]+)\s*(.*)", str(range_val))
            if range_match:
                range_min = range_match.group(1)
                range_max = range_match.group(2)
                if not range_unit:
                    range_unit = range_match.group(3).strip()

        # Location — check top-level area and location fields
        area = tag.get("area", inst.get("area", ""))
        pid_ref = inst.get("pid_reference", inst.get("source_pid", ""))
        location_str = inst.get("location", "")
        if isinstance(location_str, dict):
            pid_ref = pid_ref or location_str.get("pid_reference", "")
            location_str = location_str.get("physical_location", "")

        data = [
            item_no,
            tag.get("full_tag", ""),
            inst.get("service_description", inst.get("service", "")),
            pid_ref,
            inst.get("equipment_tag", ""),
            location_str if isinstance(location_str, str) else str(area),
            inst.get("manufacturer", ""),
            inst.get("instrument_type", inst.get("type", "")),
            signal_type,
            range_unit,
            range_min,
            range_max,
            inst.get("alarm_lolo", ""),
            inst.get("alarm_lo", ""),
            inst.get("alarm_hi", ""),
            inst.get("alarm_hihi", ""),
            range_min,  # PLC 4mA = range_min
            range_max,  # PLC 20mA = range_max
            inst.get("remarks", ""),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value if value else "")
            cell.border = thin_border
            if col == 1:  # Item number
                cell.alignment = center_align
            elif col in [10, 11, 12, 13, 14, 15, 16, 17, 18]:  # Numeric
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Freeze header
    ws.freeze_panes = "A5"

    return wb


def main():
    parser = argparse.ArgumentParser(description="Generate Instrument Index Excel")
    parser.add_argument("--database", "-d", required=True, help="Path to database YAML")
    parser.add_argument("--output", "-o", help="Output Excel path")
    args = parser.parse_args()

    database_path = Path(args.database)
    if not database_path.exists():
        print(f"Error: Database file not found: {database_path}")
        sys.exit(1)

    # Load database
    print(f"Loading database: {database_path}")
    database = load_yaml(database_path)

    instruments = database.get("instruments", [])
    project_id = database.get("project_id", "UNKNOWN")
    revision = database.get("revision", {})

    print(f"Found {len(instruments)} instruments")

    # Generate output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = database_path.parent / "instrument-index.xlsx"

    # Create workbook
    wb = create_workbook(instruments, project_id, revision)

    # Save
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
