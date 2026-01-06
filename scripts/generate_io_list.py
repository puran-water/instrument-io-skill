#!/usr/bin/env python3
"""
Generate IO List Excel from database.

Creates a 17-column Excel file matching standard IO list format.

Usage:
    python generate_io_list.py --database database.yaml
    python generate_io_list.py -d database.yaml -o io-list.xlsx
"""

import argparse
import sys
from pathlib import Path

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_yaml(path: Path) -> dict:
    """Load YAML file."""
    with open(path) as f:
        return yaml.safe_load(f)


# Column definitions for IO List
COLUMNS = [
    ("Area", 8),
    ("ISA PLC", 10),
    ("ISA Field", 10),
    ("S.No.", 8),
    ("PLC Tag Number", 18),
    ("Field Tag Number", 18),
    ("Service Description", 35),
    ("Component Description", 25),
    ("P&ID Number", 12),
    ("Instrument Location", 15),
    ("Signal Type", 12),
    ("I/O Type", 8),
    ("Signal Type (Voltage)", 15),
    ("Termination Point", 12),
    ("Function", 12),
    ("Feeder Type", 12),
    ("IO Pattern", 12),
    ("Remarks", 25),
]


def get_signal_category(io_type: str) -> str:
    """Return 'Digital' or 'Analog' based on io_type."""
    if io_type in ["DI", "DO"]:
        return "Digital"
    elif io_type in ["AI", "AO"]:
        return "Analog"
    elif io_type in ["PI", "PO"]:
        return "Protocol"
    return ""


def create_workbook(instruments: list, project_id: str, revision: dict) -> Workbook:
    """
    Create IO List workbook.

    Args:
        instruments: List of instrument entries
        project_id: Project identifier
        revision: Revision info

    Returns:
        openpyxl Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IO List"

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title row (18 columns now)
    ws.merge_cells("A1:R1")
    ws["A1"] = f"IO LIST - {project_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Revision row
    ws.merge_cells("A2:R2")
    ws["A2"] = f"Revision: {revision.get('number', 'A')} | Date: {revision.get('date', '')} | By: {revision.get('by', '')}"
    ws["A2"].alignment = center_align

    # Header row (row 4)
    header_row = 4
    for col, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows - one row per IO signal
    row_num = header_row
    item_no = 0

    for inst in instruments:
        tag = inst.get("tag", {})
        location = inst.get("location", {})
        # Handle location as string or dict
        if isinstance(location, str):
            location = {"physical_location": location, "pid_reference": ""}
        io_signals = inst.get("io_signals", [])

        if not io_signals:
            continue

        for signal in io_signals:
            item_no += 1
            row_num += 1

            io_type = signal.get("io_type", "")
            signal_category = get_signal_category(io_type)

            # Get feeder type from electrical object
            electrical = signal.get("electrical", {})
            feeder_type = electrical.get("feeder_type", "") if electrical else ""

            # Build full tag with suffix if available
            suffix = signal.get("suffix", "")
            full_tag = tag.get("full_tag", "")
            field_tag = f"{full_tag}-{suffix}" if suffix else full_tag

            # Get service description (handle both field names)
            service_desc = inst.get("service_description", "") or inst.get("service", "")

            # Get signal function (handle both field names)
            sig_function = signal.get("signal_function", "") or signal.get("function", "")

            # Get component description
            component_desc = signal.get("component_type", "") or signal.get("component", "") or inst.get("component", "")

            # Get feeder type from instrument level if not in signal
            if not feeder_type:
                feeder_type = inst.get("feeder_type", "")

            data = [
                tag.get("area", ""),
                tag.get("variable", ""),  # ISA PLC (variable letter)
                tag.get("variable", ""),  # ISA Field
                item_no,
                field_tag,  # PLC Tag Number
                field_tag,  # Field Tag Number
                service_desc,
                component_desc,
                location.get("pid_reference", ""),
                location.get("physical_location", ""),
                signal_category,
                io_type,
                signal.get("signal_type", ""),
                signal.get("termination", ""),
                sig_function,
                feeder_type,  # Electrical feeder type (DOL, VFD, etc.)
                signal.get("pattern_source", ""),  # IO pattern used
                signal.get("description", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value if value else "")
                cell.border = thin_border
                if col in [1, 2, 3, 4, 11, 12]:  # Center-aligned
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

    # Freeze header
    ws.freeze_panes = "A5"

    return wb


def main():
    parser = argparse.ArgumentParser(description="Generate IO List Excel")
    parser.add_argument("--database", "-d", required=True, help="Path to database YAML")
    parser.add_argument("--output", "-o", help="Output Excel path")
    args = parser.parse_args()

    database_path = Path(args.database)
    if not database_path.exists():
        print(f"Error: Database file not found: {database_path}")
        sys.exit(1)

    # Load database
    print(f"Loading database: {database_path}")
    database = load_yaml(database_path)

    instruments = database.get("instruments", [])
    project_id = database.get("project_id", "UNKNOWN")
    revision = database.get("revision", {})

    # Count IO signals
    total_signals = sum(len(inst.get("io_signals", [])) for inst in instruments)
    print(f"Found {len(instruments)} instruments with {total_signals} IO signals")

    # Generate output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = database_path.parent / "io-list.xlsx"

    # Create workbook
    wb = create_workbook(instruments, project_id, revision)

    # Save
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
