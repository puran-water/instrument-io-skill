#!/usr/bin/env python3
"""
Generate IO Summary Excel from database.

Counts DI/DO/AI/AO/PI/PO and calculates spare capacity.

Usage:
    python generate_io_summary.py --database database.yaml
    python generate_io_summary.py -d database.yaml --spare-pct 20 -o io-summary.xlsx
"""

import argparse
import math
import sys
from pathlib import Path

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def load_yaml(path: Path) -> dict:
    """Load YAML file."""
    with open(path) as f:
        return yaml.safe_load(f)


def count_io_types(instruments: list) -> dict:
    """
    Count IO points by type.

    Args:
        instruments: List of instrument entries

    Returns:
        Dict with counts for each IO type
    """
    counts = {"DI": 0, "DO": 0, "AI": 0, "AO": 0, "PI": 0, "PO": 0}

    for inst in instruments:
        for signal in (inst.get("io_signals") or []):
            io_type = signal.get("io_type", "")
            if io_type in counts:
                counts[io_type] += 1

    return counts


def create_workbook(counts: dict, spare_pct: float, project_id: str, revision: dict) -> Workbook:
    """
    Create IO Summary workbook.

    Args:
        counts: IO type counts
        spare_pct: Spare percentage (e.g., 20 for 20%)
        project_id: Project identifier
        revision: Revision info

    Returns:
        openpyxl Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "IO Summary"

    # Styles
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"IO SUMMARY - {project_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Revision row
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Revision: {revision.get('number', 'A')} | Date: {revision.get('date', '')} | Spare: {spare_pct}%"
    ws["A2"].alignment = center_align

    # Headers (row 4)
    headers = [
        ("IO Type", 12),
        ("Required", 12),
        (f"Spare ({spare_pct}%)", 12),
        ("Total Recommended", 16),
        ("Provided", 12),
        ("Available", 12),
    ]

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    io_types = ["DI", "DO", "AI", "AO", "PI", "PO"]
    row = 5

    for io_type in io_types:
        required = counts.get(io_type, 0)
        spare = math.ceil(required * spare_pct / 100)
        total = required + spare
        provided = ""  # User fills in
        available = ""  # User fills in

        data = [io_type, required, spare, total, provided, available]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value if value != "" else "")
            cell.border = thin_border
            if col == 1:
                cell.alignment = center_align
                cell.font = Font(bold=True)
            else:
                cell.alignment = right_align

        row += 1

    # Totals row
    total_required = sum(counts.values())
    total_spare = math.ceil(total_required * spare_pct / 100)
    total_total = total_required + total_spare

    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=total_required)
    ws.cell(row=row, column=3, value=total_spare)
    ws.cell(row=row, column=4, value=total_total)

    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.font = Font(bold=True)
        if col == 1:
            cell.alignment = center_align
        else:
            cell.alignment = right_align

    # Protocol IO note
    row += 2
    pi_count = counts.get("PI", 0)
    po_count = counts.get("PO", 0)
    if pi_count or po_count:
        ws.merge_cells(f"A{row}:F{row}")
        ws[f"A{row}"] = f"Note: Protocol IO (PI/PO) represents Modbus/HART/Profibus device connections"
        ws[f"A{row}"].font = Font(italic=True, size=9)

    return wb


def main():
    parser = argparse.ArgumentParser(description="Generate IO Summary Excel")
    parser.add_argument("--database", "-d", required=True, help="Path to database YAML")
    parser.add_argument("--spare-pct", "-s", type=float, default=20.0, help="Spare percentage (default: 20)")
    parser.add_argument("--output", "-o", help="Output Excel path")
    args = parser.parse_args()

    database_path = Path(args.database)
    if not database_path.exists():
        print(f"Error: Database file not found: {database_path}")
        sys.exit(1)

    # Load database
    print(f"Loading database: {database_path}")
    database = load_yaml(database_path)

    instruments = database.get("instruments", [])
    project_id = database.get("project_id", "UNKNOWN")
    revision = database.get("revision", {})

    # Count IO types
    counts = count_io_types(instruments)

    print(f"\nIO Count Summary:")
    for io_type, count in counts.items():
        print(f"  {io_type}: {count}")
    print(f"  Total: {sum(counts.values())}")
    print(f"  Spare: {args.spare_pct}%")

    # Generate output path
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = database_path.parent / "io-summary.xlsx"

    # Create workbook
    wb = create_workbook(counts, args.spare_pct, project_id, revision)

    # Save
    wb.save(output_path)
    print(f"\nSaved: {output_path}")


if __name__ == "__main__":
    main()
